import streamlit as st
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io
import zipfile

# --- PRECIOS BASE POR PERFIL Y TIPO DE EXAMEN ---
BASE_PRICE_TABLE = {
    "PERFIL I": {"Pre-ocupacional": 95.00, "Anual": 85.00, "De Retiro": 22.00},
    "PERFIL II": {"Pre-ocupacional": 137.00, "Anual": 112.00, "De Retiro": 34.00},
    "PERFIL III": {"Pre-ocupacional": 164.00, "Anual": 139.00, "De Retiro": 34.00},
    "PERFIL IV": {"Pre-ocupacional": 137.00, "Anual": 112.00, "De Retiro": 34.00},
    "PERFIL V": {"Pre-ocupacional": 136.00, "Anual": 111.00, "De Retiro": 34.00},
    "PERFIL VI": {"Pre-ocupacional": 151.00, "Anual": 114.00, "De Retiro": 22.00},
}


PAQUETE_BASE_PRICES = {
    "PAQUETE 1": 25.00,
    "PAQUETE 2": 35.00,
    "PAQUETE 3": 25.00,
}

CONDITIONAL_PRICES = {"ecg": 15.00, "hcg": 15.00}

FILTRO_EXAM_PRICES = {"cocaina": 12.50, "marihuana": 12.50, "audimetria": 17.00}

PROFILES_1_TO_5 = {"PERFIL I", "PERFIL II", "PERFIL III", "PERFIL IV", "PERFIL V"}

TIPO_EXAMEN_MAP = {
    "PREOCUPACIONAL": "Pre-ocupacional",
    "PRE-OCUPACIONAL": "Pre-ocupacional",
    "PRE OCUPACIONAL": "Pre-ocupacional",
    "PERIODICO": "Anual",
    "PERIÓDICO": "Anual",
    "ANUAL": "Anual",
    "RETIRO": "De Retiro",
    "DE RETIRO": "De Retiro",
    "REINCORPORACION": "Reincorporación",
    "REINCORPORACIÓN": "Reincorporación",
}

NEW_COL_HEADERS = [
    "PRECIO BASE", "COCAINA", "MARIHUANA", "TRIAJE", "AUDIMETRIA",
    "PAQUETE 1 - REINCORPORACION", "PAQUETE 2 - REINCORPORACION",
    "PAQUETE 3 - REINCORPORACION", "SUB UNIDAD BETA CUALITATIVA - MUJERES",
    "CONDICIONAL - ELECTROCARDIOGRAMA",
]

NUM_NEW_COLS = 10

SRC_COL_COCAINA = 35
SRC_COL_MARIHUANA = 36
SRC_COL_TRIAJE = 43
SRC_COL_AUDIOMETRIA = 28
SRC_COL_KLOCKOFF = 42
SRC_COL_ECG = 27
SRC_COL_HCG = 44


def normalize_tipo(tipo_raw):
    if tipo_raw is None:
        return None
    return TIPO_EXAMEN_MAP.get(str(tipo_raw).strip().upper(), None)


def _exam_agendado(val):
    """True si el examen está agendado (0=no realizado, 1=realizado; ambos cuentan)."""
    return val in (0, 1)


def _get_base_profile_key(perfil_str):
    if "FILTRO" in perfil_str:
        return "FILTRO"
    if "PAQUETE 1" in perfil_str:
        return "PAQUETE 1"
    if "PAQUETE 2" in perfil_str:
        return "PAQUETE 2"
    if "PAQUETE 3" in perfil_str:
        return "PAQUETE 3"
    if "PERFIL VI" in perfil_str:
        return "PERFIL VI"
    if "PERFIL V" in perfil_str:
        return "PERFIL V"
    if "PERFIL IV" in perfil_str:
        return "PERFIL IV"
    if "PERFIL III" in perfil_str:
        return "PERFIL III"
    if "PERFIL II" in perfil_str:
        return "PERFIL II"
    if "PERFIL I" in perfil_str:
        return "PERFIL I"
    return None


def get_pricing_columns(perfil, tipo_normalizado, sexo, exam_agendado):
    result = {
        "precio_base": None,
        "cocaina": None, "marihuana": None,
        "triaje": None, "audimetria": None,
        "paquete1": None, "paquete2": None, "paquete3": None,
        "hcg": None, "ecg": None,
    }

    if perfil is None or tipo_normalizado is None:
        return result

    perfil_str = str(perfil).strip().upper()
    is_female = str(sexo).strip().upper() in ("F", "FEMENINO", "MUJER") if sexo else False
    base_key = _get_base_profile_key(perfil_str)

    # --- PRECIO BASE ---
    if base_key == "FILTRO":
        result["precio_base"] = "-"
    elif base_key in PAQUETE_BASE_PRICES:
        result["precio_base"] = PAQUETE_BASE_PRICES[base_key]
    elif base_key and base_key in BASE_PRICE_TABLE:
        result["precio_base"] = BASE_PRICE_TABLE[base_key].get(tipo_normalizado)

    # --- PERFIL FILTRO: cocaina, marihuana, audiometria ---
    if base_key == "FILTRO":
        if exam_agendado.get("cocaina"):
            result["cocaina"] = FILTRO_EXAM_PRICES["cocaina"]
        if exam_agendado.get("marihuana"):
            result["marihuana"] = FILTRO_EXAM_PRICES["marihuana"]
        if exam_agendado.get("audimetria"):
            result["audimetria"] = FILTRO_EXAM_PRICES["audimetria"]

    # --- PERFILES I-V: solo ECG y HCG condicionales ---
    elif base_key in PROFILES_1_TO_5:
        if exam_agendado.get("ecg"):
            result["ecg"] = CONDITIONAL_PRICES["ecg"]
        if is_female and exam_agendado.get("hcg"):
            result["hcg"] = CONDITIONAL_PRICES["hcg"]

    # --- Paquetes ---
    if base_key == "PAQUETE 1":
        result["paquete1"] = PAQUETE_BASE_PRICES["PAQUETE 1"]
    elif base_key == "PAQUETE 2":
        result["paquete2"] = PAQUETE_BASE_PRICES["PAQUETE 2"]
    elif base_key == "PAQUETE 3":
        result["paquete3"] = PAQUETE_BASE_PRICES["PAQUETE 3"]

    return result


# --- STREAMLIT APP ---
st.set_page_config(page_title="Separador de Preliquidación por Proyecto", layout="wide")
st.title("Separador de Preliquidación por Proyecto")
st.markdown("Sube un archivo de preliquidación y se generará un Excel por cada proyecto encontrado.")

uploaded_file = st.file_uploader("Sube el archivo de preliquidación (.xlsx)", type=["xlsx"])

if uploaded_file is not None:
    wb = openpyxl.load_workbook(uploaded_file)
    ws = wb.active

    src_max_col = ws.max_column
    header_row = 4
    data_start_row = 5

    last_data_row = ws.max_row
    if ws.cell(row=ws.max_row, column=1).value == "TOTAL":
        last_data_row = ws.max_row - 1

    proyecto_col = 16
    tipo_examen_col = 17
    perfil_col = 18
    sexo_col = 10

    new_max_col = src_max_col + NUM_NEW_COLS

    proyectos = {}
    for row_idx in range(data_start_row, last_data_row + 1):
        proyecto = ws.cell(row=row_idx, column=proyecto_col).value
        if proyecto is None or str(proyecto).strip() == "":
            proyecto = "sin proyecto"
        if proyecto not in proyectos:
            proyectos[proyecto] = []
        proyectos[proyecto].append(row_idx)

    st.success(f"Se encontraron **{len(proyectos)}** proyectos en el archivo.")
    st.markdown("**Proyectos:**")
    for p, rows in sorted(proyectos.items(), key=lambda x: x[0]):
        st.markdown(f"- {p} ({len(rows)} atenciones)")

    if st.button("Generar Excels por Proyecto"):
        zip_buffer = io.BytesIO()

        with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
            for proyecto_name, row_indices in sorted(proyectos.items(), key=lambda x: x[0]):
                new_wb = openpyxl.Workbook()
                new_ws = new_wb.active
                new_ws.title = ws.title[:31] if len(ws.title) > 31 else ws.title

                for col in range(1, 22):
                    col_letter = get_column_letter(col)
                    if ws.column_dimensions[col_letter].width:
                        new_ws.column_dimensions[col_letter].width = ws.column_dimensions[col_letter].width
                for col in range(22, src_max_col + 1):
                    src_letter = get_column_letter(col)
                    dst_letter = get_column_letter(col + NUM_NEW_COLS)
                    if ws.column_dimensions[src_letter].width:
                        new_ws.column_dimensions[dst_letter].width = ws.column_dimensions[src_letter].width

                new_ws.row_dimensions[1].height = 15.75

                new_ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=new_max_col - 3)
                title_cell = new_ws.cell(row=2, column=2, value=proyecto_name)
                title_cell.font = Font(name="Calibri", size=11, bold=True)
                title_cell.fill = PatternFill(start_color="FF00B050", end_color="FF00B050", fill_type="solid")
                title_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                new_ws.row_dimensions[2].height = 30.0

                new_ws.merge_cells(start_row=3, start_column=22, end_row=3, end_column=new_max_col)
                catalogo_cell = new_ws.cell(row=3, column=22, value="CATALOGO")
                catalogo_cell.font = Font(name="Calibri", size=8, color="FFFFFFFF")
                catalogo_cell.fill = PatternFill(start_color="FF00B050", end_color="FF00B050", fill_type="solid")
                catalogo_cell.alignment = Alignment(horizontal="center")

                new_ws.row_dimensions[4].height = 28.5
                header_style_font = Font(name="Calibri", size=8, bold=True, color="FFFFFFFF")
                header_style_fill = PatternFill(start_color="FF80BFFF", end_color="FF80BFFF", fill_type="solid")
                header_style_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
                header_style_border = Border(
                    left=Side(style="thin", color="FF000000"),
                    right=Side(style="thin", color="FF000000"),
                    top=Side(style="thin", color="FF000000"),
                    bottom=Side(style="thin", color="FF000000"),
                )

                for col in range(1, 22):
                    src_cell = ws.cell(row=header_row, column=col)
                    dst_cell = new_ws.cell(row=4, column=col, value=src_cell.value)
                    dst_cell.font = header_style_font
                    dst_cell.fill = header_style_fill
                    dst_cell.alignment = header_style_align
                    dst_cell.border = header_style_border

                for i, header_name in enumerate(NEW_COL_HEADERS):
                    dst_cell = new_ws.cell(row=4, column=22 + i, value=header_name)
                    dst_cell.font = header_style_font
                    dst_cell.fill = header_style_fill
                    dst_cell.alignment = header_style_align
                    dst_cell.border = header_style_border

                for col in range(22, src_max_col + 1):
                    src_cell = ws.cell(row=header_row, column=col)
                    dst_col = col + NUM_NEW_COLS
                    dst_cell = new_ws.cell(row=4, column=dst_col, value=src_cell.value)
                    dst_cell.font = header_style_font
                    dst_cell.fill = header_style_fill
                    dst_cell.alignment = header_style_align
                    dst_cell.border = header_style_border

                data_style_font = Font(name="Calibri", size=8)
                data_style_align = Alignment(horizontal="center", wrap_text=True)
                data_style_border = Border(
                    left=Side(style="thin", color="FF000000"),
                    right=Side(style="thin", color="FF000000"),
                    top=Side(style="thin", color="FF000000"),
                    bottom=Side(style="thin", color="FF000000"),
                )

                new_row = 5
                for idx, src_row_idx in enumerate(row_indices, start=1):
                    new_ws.row_dimensions[new_row].height = 17.25

                    perfil = ws.cell(row=src_row_idx, column=perfil_col).value
                    tipo_raw = ws.cell(row=src_row_idx, column=tipo_examen_col).value
                    sexo = ws.cell(row=src_row_idx, column=sexo_col).value
                    tipo_norm = normalize_tipo(tipo_raw)

                    exam_agendado = {
                        "cocaina": _exam_agendado(ws.cell(row=src_row_idx, column=SRC_COL_COCAINA).value),
                        "marihuana": _exam_agendado(ws.cell(row=src_row_idx, column=SRC_COL_MARIHUANA).value),
                        "triaje": _exam_agendado(ws.cell(row=src_row_idx, column=SRC_COL_TRIAJE).value),
                        "audimetria": (
                            _exam_agendado(ws.cell(row=src_row_idx, column=SRC_COL_AUDIOMETRIA).value)
                            or _exam_agendado(ws.cell(row=src_row_idx, column=SRC_COL_KLOCKOFF).value)
                        ),
                        "ecg": _exam_agendado(ws.cell(row=src_row_idx, column=SRC_COL_ECG).value),
                        "hcg": _exam_agendado(ws.cell(row=src_row_idx, column=SRC_COL_HCG).value),
                    }

                    pricing = get_pricing_columns(perfil, tipo_norm, sexo, exam_agendado)

                    for col in range(1, 20):
                        src_cell = ws.cell(row=src_row_idx, column=col)
                        value = src_cell.value
                        if col == 1:
                            value = idx
                        dst_cell = new_ws.cell(row=new_row, column=col, value=value)
                        dst_cell.font = data_style_font
                        dst_cell.alignment = data_style_align
                        dst_cell.border = data_style_border
                        if src_cell.number_format:
                            dst_cell.number_format = src_cell.number_format

                    dst_cell = new_ws.cell(row=new_row, column=20, value=f"=+U{new_row}-S{new_row}")
                    dst_cell.font = data_style_font
                    dst_cell.alignment = data_style_align
                    dst_cell.border = data_style_border
                    dst_cell.number_format = '#,##0.00'

                    dst_cell = new_ws.cell(row=new_row, column=21, value=f"=+S{new_row}*1.18")
                    dst_cell.font = data_style_font
                    dst_cell.alignment = data_style_align
                    dst_cell.border = data_style_border
                    dst_cell.number_format = '#,##0.00'

                    breakdown_values = [
                        pricing["precio_base"],
                        pricing["cocaina"], pricing["marihuana"],
                        pricing["triaje"], pricing["audimetria"],
                        pricing["paquete1"], pricing["paquete2"], pricing["paquete3"],
                        pricing["hcg"], pricing["ecg"],
                    ]
                    for i, val in enumerate(breakdown_values):
                        dst_cell = new_ws.cell(row=new_row, column=22 + i, value=val)
                        dst_cell.font = data_style_font
                        dst_cell.alignment = data_style_align
                        dst_cell.border = data_style_border
                        if val is not None and not isinstance(val, str):
                            dst_cell.number_format = '#,##0.00'

                    for col in range(22, src_max_col + 1):
                        src_cell = ws.cell(row=src_row_idx, column=col)
                        dst_col = col + NUM_NEW_COLS
                        dst_cell = new_ws.cell(row=new_row, column=dst_col, value=src_cell.value)
                        dst_cell.font = data_style_font
                        dst_cell.alignment = data_style_align
                        dst_cell.border = data_style_border
                        if src_cell.number_format:
                            dst_cell.number_format = src_cell.number_format

                    new_row += 1

                total_row = new_row
                total_label_cell = new_ws.cell(row=total_row, column=1, value="TOTAL")
                total_label_cell.font = Font(name="Calibri", size=8, bold=True)
                total_label_cell.alignment = Alignment(horizontal="center")
                total_label_cell.border = data_style_border
                for col in range(2, 22):
                    c = new_ws.cell(row=total_row, column=col)
                    c.border = data_style_border

                for col in range(19, 22):
                    col_letter = get_column_letter(col)
                    formula = f"=SUM({col_letter}5:{col_letter}{total_row - 1})"
                    dst_cell = new_ws.cell(row=total_row, column=col, value=formula)
                    dst_cell.font = Font(name="Calibri", size=8, bold=True)
                    dst_cell.alignment = Alignment(horizontal="center")
                    dst_cell.border = data_style_border
                    dst_cell.number_format = '#,##0.00'

                for col in range(22, 32):
                    col_letter = get_column_letter(col)
                    formula = f"=SUM({col_letter}5:{col_letter}{total_row - 1})"
                    dst_cell = new_ws.cell(row=total_row, column=col, value=formula)
                    dst_cell.font = Font(name="Calibri", size=8, bold=True)
                    dst_cell.alignment = Alignment(horizontal="center")
                    dst_cell.border = data_style_border
                    dst_cell.number_format = '#,##0.00'

                for col in range(32, new_max_col + 1):
                    col_letter = get_column_letter(col)
                    formula = f"=SUM({col_letter}5:{col_letter}{total_row - 1})"
                    dst_cell = new_ws.cell(row=total_row, column=col, value=formula)
                    dst_cell.font = Font(name="Calibri", size=8, bold=True)
                    dst_cell.alignment = Alignment(horizontal="center")
                    dst_cell.border = data_style_border

                file_buffer = io.BytesIO()
                new_wb.save(file_buffer)
                file_buffer.seek(0)

                safe_name = proyecto_name.replace("/", "-").replace("\\", "-").replace(":", "-")
                filename = f"Preliquidacion_{safe_name}.xlsx"
                zf.writestr(filename, file_buffer.getvalue())

        zip_buffer.seek(0)
        st.download_button(
            label="Descargar todos los Excels (ZIP)",
            data=zip_buffer,
            file_name="Preliquidaciones_por_Proyecto.zip",
            mime="application/zip",
        )
        st.success("Archivos generados exitosamente.")
